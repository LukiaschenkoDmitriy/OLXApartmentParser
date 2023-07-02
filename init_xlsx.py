from openpyxl import Workbook, load_workbook

class OLXWorkBook():
    workBook = None
    sheet = None

    mainTemplateInformation = {"A1":"Last ID", "B1": "2"}
    namesTemplate = { "A": "Title", "B": "Price",
                      "C": "Description", "D": "URL",
                      "E": "Publish Data", "F": "Czenść",
                      "G": "Number rooms", "H": "Squad meters",
                      "I": "Kind of built", "J": "Meble",
                      "K": "State", "L": "Private"}

    def __init__(self, filename = None):
        self.clear() if filename is None else self.load(filename)

    def load(self, filename):
        self.workBook = load_workbook(filename)
        self.sheet = self.workBook.active
        self.mainTemplateInformation["B1"] = self.sheet["B1"]

    def clear(self):
        self.workBook = Workbook()
        self.sheet = self.workBook.active

        self.sheet["A1"] = "Last ID"
        self.sheet["B1"] = "2"

        for i in self.namesTemplate.keys():
            self.sheet[i + self.sheet["B1"].value] = self.namesTemplate[i]

    def save(self, filename):
        self.workBook.save(filename)